from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import openpyxl
import time

# -----------------------------------------
# LOAD EXCEL DATA
# -----------------------------------------
def load_excel(path):
    wb = openpyxl.load_workbook(path)
    sheet = wb.active
    return wb, sheet


# -----------------------------------------
# RUN SEARCH TEST
# -----------------------------------------
def run_search_test(keyword):
    driver.get("https://test.bishalkarki.com")
    time.sleep(2)

    # Locate the search box
    search_box = driver.find_element(By.ID, "search_query_top")
    search_box.clear()
    search_box.send_keys(keyword)
    search_box.send_keys(Keys.ENTER)

    time.sleep(3)

    # Collect all product items
    results = driver.find_elements(By.CSS_SELECTOR, ".product-name")

    # Validate search results
    for r in results:
        name = r.text.strip().lower()
        if name == "":
            continue
        if keyword.lower() in name:
            return True  # Valid item found

    return False  # No match found


# -----------------------------------------
# MAIN PROCESS
# -----------------------------------------
driver = webdriver.Chrome()
driver.maximize_window()

excel_path = "testdata.xlsx"  # Your file in project folder

wb, sheet = load_excel(excel_path)

# Add header "Result" if not present
sheet.cell(1, 3, "Result")

print("\n===== RUNNING EXCEL DATA-DRIVEN TESTS =====")

for row in range(2, sheet.max_row + 1):
    test_case = sheet.cell(row, 1).value
    keyword = sheet.cell(row, 2).value

    print(f"\nRunning {test_case} → Searching: {keyword}")

    result = run_search_test(keyword)

    # Write PASS/FAIL in column 3
    if result:
        print(f"✔ {test_case} PASSED")
        sheet.cell(row, 3, "PASS")
    else:
        print(f"❌ {test_case} FAILED")
        sheet.cell(row, 3, "FAIL")

# Save results to new file
output_file = "testdata_result.xlsx"
wb.save(output_file)

driver.quit()

print("\n🎉 Testing Completed!")
print(f"📄 Result file saved as: {output_file}")
